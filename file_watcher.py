import os
import time
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re

class FileWatcher:
    def __init__(self, directory, tracker_path="Tracker.xlsx"):
        self.directory = directory
        self.files = {}  # hadi list katstocker lik les fichiers o lwe9t
        self.tracker_path = tracker_path
        
    def scan_directory(self):
        # Get current files
        current_files = {}
        changes_detected = False
        
        try:
            for entry in os.scandir(self.directory):
                if entry.is_file():
                    current_files[entry.name] = entry.stat().st_mtime
        except FileNotFoundError:
            print(f"Directory {self.directory} not found. Creating it...")
            os.makedirs(self.directory, exist_ok=True)
            return False
            
        # Check for new and modified files
        for filename, mtime in current_files.items():
            if filename not in self.files:
                print(f"File created: {filename}")
                # Process new RFC files
                if "RFC" in filename and (filename.endswith('.xlsx') or filename.endswith('.csv')):
                    self.process_rfc_file(os.path.join(self.directory, filename))
                changes_detected = True
            elif mtime != self.files[filename]:
                print(f"File modified: {filename}")
                changes_detected = True
                
        # Check for deleted files
        for filename in list(self.files.keys()):
            if filename not in current_files:
                print(f"File deleted: {filename}")
                changes_detected = True
                
        # Update our file list
        self.files = current_files
        
        # Flush stdout to ensure immediate output
        if changes_detected:
            sys.stdout.flush()
            
        return changes_detected
    
    def process_rfc_file(self, file_path):
        try:
            print(f"Processing RFC file: {file_path}")
            
            # Load the RFC Excel file
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Extract required data
            project_code = self.get_cell_value(sheet, 'B12')
            description = self.get_cell_value(sheet, 'B20')
            
            # Get the request number from cell F3 and remove leading zeros
            request_number = self.get_cell_value(sheet, 'F3')
            if request_number.isdigit():
                request_number = str(int(request_number))  # Remove leading zeros
            
            # Create the request of change text
            request_of_change = f"request of change {request_number}"
            
            # Determine request nature by checking which cell has light green fill
            request_nature = ""
            for cell, nature in [('C3', 'STANDARDS EXCEPTION REQUEST'), 
                                ('C4', 'PROJECT SCOPE CHANGE REQUEST'), 
                                ('C5', 'PROCEDURE / CONTRACTUAL CHANGE REQUEST')]:
                cell_obj = sheet[cell]
                if cell_obj.fill.start_color.index == 'FFB6D7A8' or cell_obj.fill.start_color.rgb == 'FFB6D7A8':
                    request_nature = nature
                    break
            
            if not request_nature:
                # Fallback: try to read the value directly from cells
                for cell, nature in [('C3', 'STANDARDS EXCEPTION REQUEST'), 
                                    ('C4', 'PROJECT SCOPE CHANGE REQUEST'), 
                                    ('C5', 'PROCEDURE / CONTRACTUAL CHANGE REQUEST')]:
                    if sheet[cell].value == 'R' or sheet[cell].value == '✓':
                        request_nature = nature
                        break
            
            # Update the tracker file
            self.update_tracker(project_code, description, request_of_change, request_nature)
            
            print(f"Successfully processed RFC file: {file_path}")
            sys.stdout.flush()
            
        except Exception as e:
            print(f"Error processing RFC file {file_path}: {str(e)}")
            sys.stdout.flush()
    
    def get_cell_value(self, sheet, cell_ref):
        """Safely get cell value, returning empty string if cell is None"""
        cell_value = sheet[cell_ref].value
        return str(cell_value) if cell_value is not None else ""
    
    def update_tracker(self, project_code, description, request_of_change, request_nature):
        try:
            print(f"Updating tracker with: Project={project_code}, Description={description}, Request={request_of_change}")
            
            # Load the tracker workbook
            wb = openpyxl.load_workbook(self.tracker_path)
            sheet = wb.active
            
            # Find the row where the project code exists
            project_row = None
            for row in range(1, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == project_code:
                    project_row = row
                    break
            
            # If project not found, find the last row and add it
            if project_row is None:
                for row in range(sheet.max_row, 0, -1):
                    if sheet.cell(row=row, column=1).value:
                        project_row = row + 1
                        sheet.cell(row=project_row, column=1).value = project_code
                        break
            
            # Find the next empty row after the project
            insert_row = project_row + 1
            while sheet.cell(row=insert_row, column=1).value:
                insert_row += 1
            
            # Insert a new row
            sheet.insert_rows(insert_row)
            
            # Add the data with the correct mapping:
            # Column A (1): Leave empty
            # Column B (2): "PENDING"
            # Column C (3): "request of change X" (from F3 in the RFC file)
            # Column D (4): Description (from B20 in the RFC file)
            # Column E (5): "AMUN"
            
            # Leave column A empty (don't set any value)
            sheet.cell(row=insert_row, column=2).value = "PENDING"           # Status (B)
            sheet.cell(row=insert_row, column=3).value = request_of_change   # Request of change (C)
            sheet.cell(row=insert_row, column=4).value = description         # Description (D)
            sheet.cell(row=insert_row, column=5).value = "AMUN"              # Responsible (E)
            
            # Save the workbook
            wb.save(self.tracker_path)
            print(f"Tracker updated successfully at row {insert_row}")
            sys.stdout.flush()
            
        except Exception as e:
            print(f"Error updating tracker: {str(e)}")
            sys.stdout.flush()
        
    def watch(self, interval=1.0):
        print(f"Watching directory: {self.directory}")
        print(f"Using tracker file: {self.tracker_path}")
        print("Press Ctrl+C to stop")
        sys.stdout.flush()  # Ensure the initial messages are displayed immediately
        
        # Initial scan
        self.scan_directory()
        
        try:
            while True:
                time.sleep(interval)
                self.scan_directory()
        except KeyboardInterrupt:
            print("\nWatcher stopped.")

if __name__ == "__main__":
    # Use command line argument for directory or default to "./watched_folder"
    watch_dir = sys.argv[1] if len(sys.argv) > 1 else "./watched_folder"
    
    # Use second command line argument for tracker file or default
    tracker_file = sys.argv[2] if len(sys.argv) > 2 else "Tracker.xlsx"
    
    # Create the directory if it doesn't exist
    Path(watch_dir).mkdir(exist_ok=True)
    
    # Set stdout to unbuffered mode
    if sys.version_info.major >= 3:
        if hasattr(sys.stdout, 'reconfigure'):
            # Python 3.7+
            sys.stdout.reconfigure(line_buffering=True)
        else:
            # Python 3.6 or earlier
            import io
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, line_buffering=True)
    else:
        # Python 2.x
        sys.stdout = os.fdopen(sys.stdout.fileno(), 'w', 0)
    
    watcher = FileWatcher(watch_dir, tracker_file)
    watcher.watch()



